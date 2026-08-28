"""
One-off ETL script: Import real Odisha (Khordha) land-record data from bhoomirashi.gov.in export.

Source: [bhoomirashi_gov_in_auth_revamp_sdet1_cshtml_project_id_54637]_features.xlsx
  - Document Information sheet: S.O. 1988E, publish date 22/06/2020
  - Land Details: 249 rows (survey numbers in Khordha, Odisha)
  - Land Parties: 478 rows (owner/party records linked to survey numbers)

Run: python -m app.scripts.import_bhoomirashi_xlsx
"""

import asyncio
import re
import sys
import uuid
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
from geoalchemy2.shape import from_shape
from shapely.geometry import Point

from app.core.config import settings
from app.core.security import get_password_hash
from app.models.state import State, District, Village
from app.models.land import LandParcel, LandOwner, LandType, OwnershipStatus
from app.models.user import User, Role
from app.models.project import Ministry, ProjectCategory, Project, Milestone, STAGES
from app.db.base import Base


# ---------------------------------------------------------------------------
# Village coordinates — sourced from OpenStreetMap Nominatim (August 2026).
# Kanjiama, Taratua, Kumbharabasta, Wilkisannagar not resolved by Nominatim;
# coordinates estimated from Khordha tahsil centroid + census village offsets.
# ---------------------------------------------------------------------------
VILLAGE_COORDS = {
    "Kanjiama": (20.1750, 85.6300),
    "Saradhapur": (20.1860, 85.5892),  # Nominatim confirmed
    "Taratua": (20.1900, 85.6100),
    "Kumbharabasta": (20.1800, 85.6000),
    "Wilkisannagar": (20.1920, 85.6350),
    "Gurujanga": (20.1966, 85.6220),  # Nominatim confirmed
}

# Khordha district center (for map default view)
KHORDHA_CENTER = (20.1863, 85.6226)

# Optional: village boundary GeoJSON centroids (same as above, for future use)
VILLAGE_BOUNDARY_GEOJSON = None  # Extend when surveyed polygons become available


def _clean_bilingual(raw: str) -> tuple[str, str]:
    """
    Split a multi-line cell into (english, odia) components.
    The bhoomirashi format puts English text first, then Odia (Devanagari)
    on subsequent lines. Embedded quote characters are stripped.
    """
    if not raw:
        return ("", "")
    # Strip embedded quote chars from the source data
    cleaned = raw.replace('"', "").strip()
    lines = [ln.strip() for ln in cleaned.split("\n") if ln.strip()]

    english_parts: list[str] = []
    odia_parts: list[str] = []
    for line in lines:
        # Detect if a line is primarily Devanagari (U+0900–U+097F block)
        devanagari_chars = sum(1 for c in line if "\u0900" <= c <= "\u097f")
        latin_chars = sum(1 for c in line if c.isascii() and c.isalpha())
        if devanagari_chars > latin_chars:
            odia_parts.append(line)
        else:
            english_parts.append(line)

    english = " ".join(english_parts).strip()
    odia = " ".join(odia_parts).strip()
    return (english, odia)


def _parse_area(raw: str) -> Optional[float]:
    """Extract numeric hectare value from area string like '0.0607\\nHectares'."""
    if not raw:
        return None
    cleaned = raw.replace('"', "").replace("\n", " ").strip()
    # Extract the first numeric value
    match = re.search(r"(\d+\.?\d*)", cleaned)
    if match:
        return float(match.group(1))
    return None


def _clean_field(raw: str) -> str:
    """Strip quotes and whitespace from a field."""
    if not raw:
        return ""
    return raw.replace('"', "").strip()


# ---------------------------------------------------------------------------
# Geocoding cache fixture — pre-computed village centroids
# ---------------------------------------------------------------------------
def _get_village_centroid(village_name: str) -> tuple[Optional[float], Optional[float]]:
    """Get pre-computed centroid for a village."""
    return VILLAGE_COORDS.get(village_name, (None, None))


async def import_bhoomirashi(
    xlsx_path: str,
    db: AsyncSession,
    *,
    truncate: bool = False,
) -> dict:
    """
    Import bhoomirashi xlsx into the NLAMS database.

    Args:
        xlsx_path: Path to the .xlsx file
        db: Async SQLAlchemy session
        truncate: If True, delete all existing land parcels/owners in the target
                  state before re-importing (idempotent mode)

    Returns:
        Summary dict with counts
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # --- 1. Read Document Information ---
    doc_info_ws = wb["Document Information"]
    doc_info_rows = list(doc_info_ws.iter_rows(values_only=True))
    provenance = {}
    for row in doc_info_rows[1:]:
        if row[0] and row[1]:
            provenance[str(row[0]).strip()] = str(row[1]).strip()
    print(f"  📄 Provenance: {provenance}")

    # --- 2. Read Land Details ---
    land_details_ws = wb["Land Details"]
    ld_rows = list(land_details_ws.iter_rows(values_only=True))
    ld_header = [str(c).strip() if c else "" for c in ld_rows[0]]
    print(f"  📋 Land Details header: {ld_header}")

    land_details = []
    for row in ld_rows[1:]:
        sno_raw = str(row[0]).strip().strip('"')
        try:
            sno = int(sno_raw)
        except (ValueError, TypeError):
            continue

        survey_number_en, survey_number_or = _clean_bilingual(str(row[4]) if row[4] else "")
        area_raw = str(row[5]) if row[5] else ""
        area_hectares = _parse_area(area_raw)

        desc_en, desc_or = _clean_bilingual(str(row[6]) if row[6] else "")

        land_type_raw = _clean_field(str(row[7]) if row[7] else "")
        land_nature_raw = _clean_field(str(row[8]) if row[8] else "")
        land_category_raw = _clean_field(str(row[9]) if row[9] else "")

        additional_details = _clean_field(str(row[10]) if row[10] else "")

        land_details.append({
            "sno": sno,
            "district": str(row[1]).strip().strip('"') if row[1] else "",
            "sub_district": str(row[2]).strip().strip('"') if row[2] else "",
            "village": str(row[3]).strip().strip('"') if row[3] else "",
            "survey_number": sno_raw,
            "survey_number_en": survey_number_en,
            "survey_number_or": survey_number_or,
            "area_hectares": area_hectares,
            "area_raw": area_raw.replace('"', "").strip(),
            "description_en": desc_en,
            "description_or": desc_or,
            "land_type_raw": land_type_raw,
            "land_nature_raw": land_nature_raw,
            "land_category_raw": land_category_raw,
            "additional_details": additional_details,
        })

    print(f"  📊 Parsed {len(land_details)} land detail rows")

    # --- 3. Read Land Parties ---
    land_parties_ws = wb["Land Parties"]
    lp_rows = list(land_parties_ws.iter_rows(values_only=True))

    # Group by Source S.No
    parties_by_sno: dict[int, list[dict]] = {}
    for row in lp_rows[1:]:
        sno_raw = str(row[0]).strip().strip('"')
        try:
            sno = int(sno_raw)
        except (ValueError, TypeError):
            continue

        name_en, name_or = _clean_bilingual(str(row[2]) if row[2] else "")
        addr_en, addr_or = _clean_bilingual(str(row[3]) if row[3] else "")
        party_type = _clean_field(str(row[4]) if row[4] else "Owner")
        area_str = str(row[5]).replace('"', "").strip() if row[5] else ""

        party = {
            "name_en": name_en,
            "name_or": name_or,
            "address_en": addr_en,
            "address_or": addr_or,
            "type": party_type,
            "area_str": area_str,
        }
        parties_by_sno.setdefault(sno, []).append(party)

    print(f"  👥 Parsed {sum(len(v) for v in parties_by_sno.values())} party records across {len(parties_by_sno)} survey numbers")

    wb.close()

    # --- 4. Create State, District, Villages ---
    state = (await db.execute(
        select(State).where(State.code == "OD")
    )).scalar_one_or_none()
    if not state:
        state = State(name="Odisha", code="OD", region="East")
        db.add(state)
        await db.flush()
        print("  🏛️ Created State: Odisha (OD)")

    district = (await db.execute(
        select(District).where(District.state_id == state.id, District.name == "Khordha")
    )).scalar_one_or_none()
    if not district:
        district = District(state_id=state.id, name="Khordha", code="KHD")
        db.add(district)
        await db.flush()
        print("  🏛️ Created District: Khordha")

    villages_map: dict[str, Village] = {}
    for v_name in sorted(set(ld["village"] for ld in land_details)):
        existing = (await db.execute(
            select(Village).where(
                Village.district_id == district.id,
                Village.name == v_name,
            )
        )).scalar_one_or_none()
        if existing:
            villages_map[v_name] = existing
        else:
            lat, lon = _get_village_centroid(v_name)
            v = Village(
                district_id=district.id,
                tehsil="Khordha",
                name=v_name,
                code=f"KHD-{v_name[:8].upper()}",
                latitude=lat,
                longitude=lon,
            )
            db.add(v)
            await db.flush()
            villages_map[v_name] = v
            print(f"  🏘️ Created Village: {v_name} ({lat}, {lon})")

    # --- 5. Create a default project for these parcels ---
    # Use or create a single project for the bhoomirashi import
    project = (await db.execute(
        select(Project).where(Project.name == "S.O. 1988E — Khordha Land Acquisition")
    )).scalar_one_or_none()
    if not project:
        # Ensure ministry and category exist
        ministry = (await db.execute(
            select(Ministry).where(Ministry.code == "MoRTH")
        )).scalar_one_or_none()
        if not ministry:
            ministry = Ministry(name="Ministry of Road Transport & Highways", code="MoRTH")
            db.add(ministry)
            await db.flush()

        category = (await db.execute(
            select(ProjectCategory).where(ProjectCategory.name == "Highway")
        )).scalar_one_or_none()
        if not category:
            category = ProjectCategory(name="Highway")
            db.add(category)
            await db.flush()

        # Find super_admin user for created_by
        super_admin = (await db.execute(
            select(User).join(Role).where(Role.name == "super_admin").limit(1)
        )).scalar_one_or_none()

        project = Project(
            name="S.O. 1988E — Khordha Land Acquisition",
            ministry_id=ministry.id,
            category_id=category.id,
            state_id=state.id,
            district_id=district.id,
            implementing_agency_id=super_admin.id if super_admin else None,
            description="Land acquisition for S.O. 1988E notification, Khordha district, Odisha. Source: bhoomirashi.gov.in export dated 22/06/2020.",
            estimated_budget=50000000,
            estimated_land_required_hectares=sum(ld["area_hectares"] or 0 for ld in land_details),
            priority="high",
            status="active",
            current_stage="gis_mapping",
            created_by=super_admin.id if super_admin else uuid.uuid4(),
        )
        db.add(project)
        await db.flush()
        print("  📁 Created Project: S.O. 1988E — Khordha Land Acquisition")

    # --- 6. Optionally truncate existing parcels in this state ---
    if truncate:
        # Delete owners first (FK dependency)
        parcel_ids_subq = select(LandParcel.id).where(LandParcel.state_id == state.id)
        await db.execute(
            text("DELETE FROM land_owners WHERE parcel_id IN (SELECT id FROM land_parcels WHERE state_id = :sid)"),
            {"sid": state.id},
        )
        await db.execute(
            text("DELETE FROM land_parcels WHERE state_id = :sid"),
            {"sid": state.id},
        )
        await db.flush()
        print("  🗑️ Truncated existing parcels in Odisha")

    # --- 7. Upsert parcels + owners ---
    created_parcels = 0
    created_owners = 0
    skipped = 0

    for ld in land_details:
        sno = ld["sno"]
        village = villages_map.get(ld["village"])
        if not village:
            print(f"  ⚠️ Skipping S.No {sno}: village '{ld['village']}' not found")
            skipped += 1
            continue

        # Check for existing parcel (idempotent on survey_number + village)
        existing = (await db.execute(
            select(LandParcel).where(
                LandParcel.survey_number == ld["survey_number"],
                LandParcel.village_id == village.id,
                LandParcel.state_id == state.id,
            )
        )).scalar_one_or_none()

        if existing and not truncate:
            # Update existing
            parcel = existing
            parcel.area_hectares = ld["area_hectares"]
        elif not existing:
            # Map land type
            land_type = LandType.wet  # All bhoomirashi land is "Wet"
            # Map ownership status
            nature = ld["land_nature_raw"].lower()
            if "government" in nature:
                ownership = OwnershipStatus.govt
            elif "private" in nature:
                ownership = OwnershipStatus.private
            else:
                ownership = OwnershipStatus.private

            # Create Point geometry at village centroid
            lat, lon = _get_village_centroid(ld["village"])
            geom = None
            if lat is not None and lon is not None:
                # Jitter parcels slightly around village centroid to avoid stacking
                import random
                jitter_lat = lat + random.uniform(-0.002, 0.002)
                jitter_lon = lon + random.uniform(-0.002, 0.002)
                geom = from_shape(Point(jitter_lon, jitter_lat), srid=4326)

            parcel = LandParcel(
                project_id=project.id,
                survey_number=ld["survey_number"],
                village_id=village.id,
                district_id=district.id,
                state_id=state.id,
                area_hectares=ld["area_hectares"],
                geom=geom,
                land_type=land_type,
                ownership_status=ownership,
            )
            db.add(parcel)
            created_parcels += 1
        else:
            skipped += 1
            continue

        await db.flush()

        # Create owners for this parcel
        parties = parties_by_sno.get(sno, [])
        if parties:
            # Remove existing owners if updating
            if existing and not truncate:
                for old_owner in list(existing.owners or []):
                    await db.delete(old_owner)
                await db.flush()

            for party in parties:
                # Some 'Name' cells contain multiple co-owners separated by commas
                # Split them into separate LandOwner records
                raw_name = party["name_en"] or party["name_or"] or "Unknown"
                # Split on comma, but only if we get meaningful parts (>3 chars each)
                parts = [p.strip() for p in raw_name.split(",") if p.strip()]
                if len(parts) > 1 and all(len(p) > 3 for p in parts):
                    names_to_create = parts
                else:
                    names_to_create = [raw_name]

                for name in names_to_create:
                    owner = LandOwner(
                        parcel_id=parcel.id,
                        full_name=name[:500],  # Safety truncation
                        phone="0000000000",  # No phone in source data
                    )
                    db.add(owner)
                    created_owners += 1

    await db.flush()
    print(f"\n  ✅ Import complete:")
    print(f"     Parcels created: {created_parcels}")
    print(f"     Owners created: {created_owners}")
    print(f"     Skipped: {skipped}")
    print(f"     State: Odisha (OD)")
    print(f"     District: Khordha")
    print(f"     Villages: {len(villages_map)}")

    return {
        "parcels_created": created_parcels,
        "owners_created": created_owners,
        "skipped": skipped,
        "state": "Odisha",
        "district": "Khordha",
        "villages": list(villages_map.keys()),
    }


async def create_tables():
    """Create all tables from metadata (for fresh DB setup)."""
    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    async with engine.begin() as conn:
        await conn.execute(text("CREATE EXTENSION IF NOT EXISTS postgis"))
        await conn.execute(text("CREATE EXTENSION IF NOT EXISTS pg_trgm"))
        await conn.run_sync(Base.metadata.create_all)
    return engine


async def main():
    """Main entry point for the import script."""
    import os

    # Find the xlsx file
    xlsx_path = None
    for candidate in [
        os.environ.get("BHOOMIRASHI_XLSX", ""),
        str(Path(__file__).parent.parent.parent.parent / "[bhoomirashi_gov_in_auth_revamp_sdet1_cshtml_project_id_54637]_features (1).xlsx"),
        str(Path(__file__).parent.parent.parent.parent / "_bhoomirashi_gov_in_auth_revamp_sdet1_cshtml_project_id_54637__features.xlsx"),
    ]:
        if candidate and Path(candidate).exists():
            xlsx_path = candidate
            break

    if not xlsx_path:
        print("❌ xlsx file not found. Set BHOOMIRASHI_XLSX env var or place the file in the repo root.")
        sys.exit(1)

    print(f"📂 Reading from: {xlsx_path}")

    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    async_session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with async_session() as db:
        result = await import_bhoomirashi(xlsx_path, db, truncate=True)
        await db.commit()

    await engine.dispose()
    print("\n🎉 Import complete!")
    return result


if __name__ == "__main__":
    asyncio.run(main())
